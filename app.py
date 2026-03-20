import streamlit as st
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import openpyxl
from collections import defaultdict, Counter
import io
import r
from datetime import date

# ══════════════════════════════════════════════════════════════════════════
# CONFIG & STYLE
# ══════════════════════════════════════════════════════════════════════════
st.set_page_config(
    page_title="Thales — Besoins & Candidatures",
    page_icon="📊", layout="wide",
    initial_sidebar_state="expanded"
)

st.markdown("""
<style>
    .main { padding-top: 0.5rem; }
    .metric-card {
        background: linear-gradient(135deg, #1F3864, #2E75B6);
        border-radius: 12px; padding: 18px; color: white;
        text-align: center; margin: 3px;
    }
    .metric-card.green  { background: linear-gradient(135deg, #375623, #4CAF50); }
    .metric-card.blue   { background: linear-gradient(135deg, #1a5276, #2E75B6); }
    .metric-card.orange { background: linear-gradient(135deg, #784212, #C55A11); }
    .metric-card.red    { background: linear-gradient(135deg, #7B0000, #C00000); }
    .metric-card.grey   { background: linear-gradient(135deg, #444, #777); }
    .metric-card h2 { font-size: 2.2rem; margin: 0; font-weight: 700; }
    .metric-card p  { margin: 4px 0 0; font-size: 0.82rem; opacity: 0.9; }
    .metric-card .pct { font-size: 0.78rem; opacity: 0.75; }
    .search-box { background: #f0f4f8; border-radius: 10px; padding: 16px; margin-bottom: 12px; }
    .section-title { font-size: 1.1rem; font-weight: 700; color: #1F3864;
                     border-left: 4px solid #2E75B6; padding-left: 10px; margin: 16px 0 8px; }
    div[data-testid="stSidebarNav"] { display: none; }
    .stTabs [data-baseweb="tab-list"] { gap: 4px; }
    .stTabs [data-baseweb="tab"] { padding: 6px 16px; border-radius: 6px 6px 0 0; }
</style>
""", unsafe_allow_html=True)

# ══════════════════════════════════════════════════════════════════════════
# CONSTANTES
# ══════════════════════════════════════════════════════════════════════════
CAT_ORDER  = ['acceptee','a_selectionner','a_etudier','toutes_refusees','sans_cand']
CAT_LBL    = {
    'acceptee':       '✅ Acceptée',
    'a_selectionner': '🔵 A sélectionner',
    'a_etudier':      '🔵 A étudier',
    'toutes_refusees':'🟠 Toutes refusées',
    'sans_cand':      '🔴 Sans candidature',
}
CAT_COLORS = {
    'acceptee':       '#375623',
    'a_selectionner': '#2E75B6',
    'a_etudier':      '#5BA3E0',
    'toutes_refusees':'#C55A11',
    'sans_cand':      '#C00000',
}
CAT_PRIO = {'acceptee':0,'a_selectionner':1,'a_etudier':2,'toutes_refusees':3,'sans_cand':4}

# ══════════════════════════════════════════════════════════════════════════
# CHARGEMENT FICHIERS (xlsx / xls / csv)
# ══════════════════════════════════════════════════════════════════════════
def lire_fichier(uploaded_file):
    name = uploaded_file.name.lower()
    if name.endswith('.csv'):
        import chardet
        raw = uploaded_file.read()
        encoding = chardet.detect(raw)['encoding'] or 'utf-8'
        try:
            text = raw.decode(encoding)
        except Exception:
            text = raw.decode('latin-1')
        buf = io.StringIO(text)
        try:
            df = pd.read_csv(buf, sep=None, engine='python', dtype=str, keep_default_na=False)
        except Exception:
            buf.seek(0)
            df = pd.read_csv(buf, sep=';', dtype=str, keep_default_na=False)
        return [tuple(df.columns.tolist())] + [tuple(r) for r in df.itertuples(index=False, name=None)]
    elif name.endswith('.xls'):
        import xlrd
        uploaded_file.seek(0)
        wb = xlrd.open_workbook(file_contents=uploaded_file.read())
        ws = wb.sheet_by_index(0)
        return [tuple(ws.row_values(i)) for i in range(ws.nrows)]
    else:
        uploaded_file.seek(0)
        wb = openpyxl.load_workbook(uploaded_file, data_only=True)
        return list(wb.active.iter_rows(values_only=True))

def find_col(headers, keywords, default):
    for i, h in enumerate(headers):
        if any(k.lower() in str(h).lower() for k in keywords):
            return i
    return default

def load_edb(uploaded_file, idx_edb_hint=None):
    rows = lire_fichier(uploaded_file)
    hdrs = [str(h).strip() if h else '' for h in rows[0]]

    idx_num  = find_col(hdrs, ['Numéro Expression', 'EdB - Numéro'], 9)
    idx_site = find_col(hdrs, ['Site'], 2)
    idx_qual = find_col(hdrs, ['Qualification'], 7)
    idx_sem  = find_col(hdrs, ['Date de diffusion', 'Semaines dans Date de diffusion'], 5)
    idx_deb  = find_col(hdrs, ['Date de début', 'Semaines dans Date de début'], 6)
    idx_ag   = find_col(hdrs, ['Agence'], 3)
    idx_siret= find_col(hdrs, ['SIRET'], 4)
    idx_sc   = find_col(hdrs, ['Candidature - Statut'], 10)
    idx_edb  = find_col(hdrs, ['Statut Expression', 'EdB - Statut'], idx_edb_hint) if idx_edb_hint is not None else None
    idx_nb   = find_col(hdrs, ['NB_CANDIDATURES'], 14)
    idx_acc  = find_col(hdrs, ['Candidatures Acceptées'], 19)
    idx_del  = find_col(hdrs, ['Diffusion / Début', 'Diffusion / Debut', 'Exprimé En Jours'], 12)

    data = [r for r in rows[1:] if r[0] and str(r[0]).strip() not in ('', 'Total général')]

    edb_d = {}
    for r in data:
        def get(i):
            try: return r[i] if i is not None and i < len(r) else None
            except: return None

        num = get(idx_num)
        if not num or str(num).strip() in ('', 'N\\A', 'N/A'): continue
        num = str(num).strip()

        statut = str(get(idx_edb)).strip() if idx_edb is not None and get(idx_edb) else None

        if num not in edb_d:
            edb_d[num] = {
                'statut': statut, 'sem': str(get(idx_sem) or '').strip(),
                'sem_debut': str(get(idx_deb) or '').strip(),
                'site': str(get(idx_site) or '').strip(),
                'qual': str(get(idx_qual) or '').strip(),
                'siret': str(get(idx_siret) or '').strip(),
                'nb_cand': 0, 'acceptee_ci': 0,
                'statuts_cand': set(), 'agences': set(),
                'delai': None,
            }
        try:    nb  = float(get(idx_nb)  or 0)
        except: nb  = 0
        try:    acc = float(get(idx_acc) or 0)
        except: acc = 0
        try:
            d = get(idx_del)
            if d is not None and str(d).strip() not in ('', 'None'):
                edb_d[num]['delai'] = float(d)
        except: pass

        edb_d[num]['nb_cand']     = max(edb_d[num]['nb_cand'],     nb)
        edb_d[num]['acceptee_ci'] = max(edb_d[num]['acceptee_ci'], acc)

        sc = get(idx_sc)
        if sc and str(sc).strip() not in ('-', '', 'nan'):
            edb_d[num]['statuts_cand'].add(str(sc).strip())
        ag = get(idx_ag)
        if ag and str(ag).strip():
            edb_d[num]['agences'].add(str(ag).strip())

    return edb_d

def categorise(e):
    sc = e['statuts_cand']; nb = e['nb_cand']
    if nb == 0:                    return 'sans_cand'
    elif 'Acceptée' in sc:         return 'acceptee'
    elif 'A sélectionner' in sc:   return 'a_selectionner'
    elif 'A étudier' in sc:        return 'a_etudier'
    else:                          return 'toutes_refusees'

def edb_to_df(edb_d, fournisseur):
    rows = []
    for num, e in edb_d.items():
        e['cat'] = categorise(e)
        # Extraire division Thales (ex : "TAS" depuis "TAS Toulouse")
        site = e['site'] or ''
        division = site.split()[0] if site else ''
        rows.append({
            'num': num, 'site': site, 'division': division,
            'qual': e['qual'], 'sem': e['sem'], 'sem_debut': e['sem_debut'],
            'statut_edb': e['statut'] or '', 'nb_cand': e['nb_cand'],
            'siret': e['siret'],
            'agences': '; '.join(sorted(e['agences'])),
            'nb_agences': len(e['agences']),
            'delai': e['delai'],
            'cat': e['cat'], 'cat_label': CAT_LBL[e['cat']],
            'fournisseur': fournisseur,
        })
    return pd.DataFrame(rows)

def consolider(df_ri, df_exp):
    all_nums = set(df_ri['num'].tolist()) | set(df_exp['num'].tolist())
    ri_map  = df_ri.set_index('num').to_dict('index')
    exp_map = df_exp.set_index('num').to_dict('index')
    rows = []
    for num in all_nums:
        ri = ri_map.get(num); ex = exp_map.get(num); base = ex if ex else ri
        cat_ri = ri['cat'] if ri else 'sans_cand'
        cat_ex = ex['cat'] if ex else 'sans_cand'
        cat_best = cat_ri if CAT_PRIO[cat_ri] <= CAT_PRIO[cat_ex] else cat_ex
        delai = ex['delai'] if ex and ex['delai'] is not None else (ri['delai'] if ri else None)
        rows.append({
            'num': num, 'site': base['site'], 'division': base['division'],
            'qual': base['qual'], 'sem': base['sem'], 'sem_debut': base['sem_debut'],
            'statut_edb': ex['statut_edb'] if ex else '',
            'siret': base['siret'],
            'agences': (ex['agences'] if ex else '') + (' | ' + ri['agences'] if ri and ri['agences'] else ''),
            'nb_cand_ri': ri['nb_cand'] if ri else 0,
            'nb_cand_exp': ex['nb_cand'] if ex else 0,
            'nb_cand': (ri['nb_cand'] if ri else 0) + (ex['nb_cand'] if ex else 0),
            'delai': delai,
            'cat_ri': cat_ri, 'cat_exp': cat_ex, 'cat': cat_best,
            'cat_label': CAT_LBL[cat_best],
            'fournisseur': 'RI+EXP' if ri and ex else ('RI' if ri else 'EXP'),
        })
    return pd.DataFrame(rows)

# ══════════════════════════════════════════════════════════════════════════
# COMPOSANTS GRAPHIQUES
# ══════════════════════════════════════════════════════════════════════════
def kpi_row(df):
    total = len(df)
    if total == 0: return
    counts = df['cat'].value_counts()
    cols = st.columns(5)
    for i, (c_id, style) in enumerate(zip(CAT_ORDER, ['','green','blue','orange','red'])):
        val = counts.get(c_id, 0)
        with cols[i]:
            st.markdown(f"""
            <div class="metric-card {style}">
                <p>{CAT_LBL[c_id]}</p>
                <h2>{val}</h2>
                <p class="pct">{val/total*100:.0f}% du total</p>
            </div>""", unsafe_allow_html=True)

def fig_camembert(df, titre):
    counts = df['cat'].value_counts().reindex(CAT_ORDER, fill_value=0)
    fig = go.Figure(go.Pie(
        labels=[CAT_LBL[c] for c in CAT_ORDER], values=counts.values, hole=0.44,
        marker_colors=[CAT_COLORS[c] for c in CAT_ORDER],
        textinfo='percent+label', textfont_size=10,
    ))
    fig.update_layout(title=titre, showlegend=False, height=310,
                      margin=dict(t=40,b=5,l=5,r=5), font=dict(family='Arial'))
    return fig

def fig_barres_semaine(df, titre):
    sem_cat = df.groupby(['sem','cat']).size().unstack(fill_value=0)
    for c in CAT_ORDER:
        if c not in sem_cat.columns: sem_cat[c] = 0
    sem_cat = sem_cat[CAT_ORDER].sort_index()
    fig = go.Figure()
    for c in CAT_ORDER:
        fig.add_trace(go.Bar(name=CAT_LBL[c], x=sem_cat.index,
                             y=sem_cat[c], marker_color=CAT_COLORS[c]))
    fig.update_layout(barmode='stack', title=titre, height=360,
                      xaxis_title='Semaine', yaxis_title='Nb Expressions',
                      legend=dict(orientation='h', y=-0.28),
                      margin=dict(t=40,b=70,l=40,r=10),
                      font=dict(family='Arial'), plot_bgcolor='#FAFAFA')
    fig.update_xaxes(tickangle=45)
    return fig

def fig_sites(df, titre, top_n=15):
    sc = df.groupby(['site','cat']).size().unstack(fill_value=0)
    for c in CAT_ORDER:
        if c not in sc.columns: sc[c] = 0
    sc['critique'] = sc.get('sans_cand', 0) + sc.get('toutes_refusees', 0)
    sc = sc.sort_values('critique', ascending=False).head(top_n)[CAT_ORDER].iloc[::-1]
    fig = go.Figure()
    for c in CAT_ORDER:
        fig.add_trace(go.Bar(name=CAT_LBL[c], y=sc.index, x=sc[c],
                             orientation='h', marker_color=CAT_COLORS[c]))
    fig.update_layout(barmode='stack', title=titre, height=max(360, top_n*26),
                      xaxis_title='Nb Expressions', legend=dict(orientation='h', y=-0.14),
                      margin=dict(t=40,b=50,l=10,r=10),
                      font=dict(family='Arial'), plot_bgcolor='#FAFAFA')
    return fig

def fig_taux_couverture(df, group_col, titre, top_n=15):
    """Taux de couverture (% avec candidature) par groupement"""
    df2 = df.copy()
    df2['avec'] = (df2['cat'] != 'sans_cand').astype(int)
    grp = df2.groupby(group_col).agg(total=('num','count'), avec=('avec','sum')).reset_index()
    grp['taux'] = grp['avec'] / grp['total'] * 100
    grp = grp.sort_values('taux').tail(top_n)
    colors = ['#C00000' if t < 40 else '#C55A11' if t < 70 else '#375623' for t in grp['taux']]
    fig = go.Figure(go.Bar(
        x=grp['taux'], y=grp[group_col], orientation='h',
        marker_color=colors,
        text=[f"{t:.0f}% ({a}/{tot})" for t,a,tot in zip(grp['taux'],grp['avec'],grp['total'])],
        textposition='outside',
    ))
    fig.update_layout(title=titre, height=max(320, top_n*26), xaxis_title='% avec candidature',
                      xaxis=dict(range=[0,115]), margin=dict(t=40,b=40,l=10,r=80),
                      font=dict(family='Arial'), plot_bgcolor='#FAFAFA')
    fig.add_vline(x=70, line_dash='dash', line_color='#2E75B6', annotation_text='Seuil 70%')
    return fig

def fig_tendance(df, titre):
    """Évolution semaine par semaine — courbes"""
    sem_cat = df.groupby(['sem','cat']).size().unstack(fill_value=0).sort_index()
    for c in CAT_ORDER:
        if c not in sem_cat.columns: sem_cat[c] = 0
    fig = go.Figure()
    for c in ['sans_cand','toutes_refusees','a_selectionner','a_etudier','acceptee']:
        fig.add_trace(go.Scatter(
            x=sem_cat.index, y=sem_cat[c], mode='lines+markers',
            name=CAT_LBL[c], line=dict(color=CAT_COLORS[c], width=2),
            marker=dict(size=6)
        ))
    fig.update_layout(title=titre, height=380, xaxis_title='Semaine',
                      yaxis_title='Nb Expressions',
                      legend=dict(orientation='h', y=-0.25),
                      margin=dict(t=40,b=70,l=40,r=10),
                      font=dict(family='Arial'), plot_bgcolor='#FAFAFA')
    fig.update_xaxes(tickangle=45)
    return fig

def fig_qualifications(df, top_n=20):
    """Top qualifications en tension"""
    q = df.groupby(['qual','cat']).size().unstack(fill_value=0)
    for c in CAT_ORDER:
        if c not in q.columns: q[c] = 0
    q['tension'] = q.get('sans_cand',0) + q.get('toutes_refusees',0)
    q = q.sort_values('tension', ascending=False).head(top_n)[CAT_ORDER].iloc[::-1]
    # Tronquer les labels trop longs
    q.index = [l[:45]+'…' if len(l)>45 else l for l in q.index]
    fig = go.Figure()
    for c in CAT_ORDER:
        fig.add_trace(go.Bar(name=CAT_LBL[c], y=q.index, x=q[c],
                             orientation='h', marker_color=CAT_COLORS[c]))
    fig.update_layout(barmode='stack', title=f"Top {top_n} Qualifications en tension",
                      height=max(380, top_n*22),
                      xaxis_title='Nb Expressions', legend=dict(orientation='h', y=-0.12),
                      margin=dict(t=40,b=50,l=10,r=10),
                      font=dict(family='Arial'), plot_bgcolor='#FAFAFA')
    return fig

def fig_delais(df, group_col, titre):
    """Délais moyens par groupement"""
    df2 = df[df['delai'].notna()].copy()
    if df2.empty:
        return None
    grp = df2.groupby(group_col)['delai'].agg(['mean','median','count']).reset_index()
    grp.columns = [group_col, 'moyenne', 'médiane', 'count']
    grp = grp[grp['count'] >= 2].sort_values('moyenne')
    if grp.empty: return None
    fig = go.Figure()
    fig.add_trace(go.Bar(name='Délai moyen (j)', x=grp[group_col], y=grp['moyenne'],
                         marker_color='#2E75B6', text=[f"{v:.0f}j" for v in grp['moyenne']],
                         textposition='outside'))
    fig.add_trace(go.Scatter(name='Médiane (j)', x=grp[group_col], y=grp['médiane'],
                             mode='markers', marker=dict(color='#C55A11', size=10, symbol='diamond')))
    fig.update_layout(title=titre, height=350, yaxis_title='Jours',
                      xaxis_tickangle=45, legend=dict(orientation='h', y=-0.3),
                      margin=dict(t=40,b=80,l=40,r=10),
                      font=dict(family='Arial'), plot_bgcolor='#FAFAFA')
    return fig

def fig_ri_vs_exp(df_ri, df_exp):
    """Comparaison RI vs EXP par site"""
    sites_communs = set(df_ri['site']) & set(df_exp['site'])
    rows = []
    for site in sites_communs:
        for fdf, fname in [(df_ri, 'Randstad Intérim'), (df_exp, 'Expectra')]:
            sub = fdf[fdf['site'] == site]
            total = len(sub)
            sans = (sub['cat'] == 'sans_cand').sum()
            acc  = (sub['cat'] == 'acceptee').sum()
            rows.append({'site': site, 'fournisseur': fname,
                         'taux_couverture': (total-sans)/total*100 if total else 0,
                         'taux_acceptation': acc/total*100 if total else 0,
                         'total': total})
    if not rows: return None
    df_cmp = pd.DataFrame(rows)
    df_cmp = df_cmp[df_cmp['total'] >= 3]
    fig = px.scatter(df_cmp, x='taux_couverture', y='taux_acceptation',
                     color='fournisseur', size='total', hover_name='site',
                     color_discrete_map={'Randstad Intérim':'#1F6B75', 'Expectra':'#7030A0'},
                     labels={'taux_couverture':'Taux couverture (%)',
                             'taux_acceptation':'Taux acceptation (%)'},
                     title='Performance RI vs EXP par site (taille = nb expressions)')
    fig.update_layout(height=420, font=dict(family='Arial'))
    return fig

def export_excel(df, nom):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    wb = Workbook(); ws = wb.active; ws.title = "Données"
    def tb():
        s = Side(style='thin', color="BFBFBF")
        return Border(left=s, right=s, top=s, bottom=s)
    cols_all = ['num','site','division','qual','sem','statut_edb','cat_label',
                'nb_cand','siret','agences','fournisseur']
    cols = [c for c in cols_all if c in df.columns]
    rename = {'num':'N° Expression','site':'Site','division':'Division','qual':'Qualification',
              'sem':'Semaine','statut_edb':'Statut EdB','cat_label':'Situation',
              'nb_cand':'Nb Cand.','siret':'SIRET','agences':'Agences','fournisseur':'Fournisseur'}
    for i, c in enumerate(cols, 1):
        cell = ws.cell(1, i, rename.get(c, c))
        cell.font = Font(name='Arial', bold=True, color='FFFFFF', size=9)
        cell.fill = PatternFill('solid', start_color='1F3864')
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = tb()
        ws.column_dimensions[get_column_letter(i)].width = 18
    ws.row_dimensions[1].height = 28
    BG = {'acceptee':'E2EFDA','a_selectionner':'D6E4F0','a_etudier':'EBF3FB',
          'toutes_refusees':'FBE4D5','sans_cand':'FCE4D6'}
    for ri, row in enumerate(df[cols].itertuples(index=False), 2):
        bg = BG.get(df.iloc[ri-2].get('cat', 'sans_cand'), 'FFFFFF')
        for ci, v in enumerate(row, 1):
            cell = ws.cell(ri, ci, str(v) if v is not None else '')
            cell.font = Font(name='Arial', size=8)
            cell.fill = PatternFill('solid', start_color=bg)
            cell.alignment = Alignment(horizontal='left', vertical='center')
            cell.border = tb()
        ws.row_dimensions[ri].height = 14
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    st.download_button(f"⬇️ Exporter Excel", buf,
                       file_name=f"{nom}_{date.today().strftime('%Y%m%d')}.xlsx",
                       mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                       use_container_width=True)

# ══════════════════════════════════════════════════════════════════════════
# SIDEBAR
# ══════════════════════════════════════════════════════════════════════════
with st.sidebar:
    try:
        st.image("https://upload.wikimedia.org/wikipedia/commons/thumb/1/16/Randstad_logo.svg/320px-Randstad_logo.svg.png", width=150)
    except: pass
    st.markdown("---")

    # ── GitHub : chargement automatique si configuré ──────────────────
    try:
        github_base = st.secrets.get("GITHUB_RAW_URL", "")
    except Exception:
        github_base = ""

    github_ri = github_exp = None
    if github_base:
        github_ri  = f"{github_base.rstrip('/')}/data/randstad_interims.xlsx"
        github_exp = f"{github_base.rstrip('/')}/data/expectra.xlsx"

    # ── Upload manuel (toujours disponible, prioritaire sur GitHub) ───
    st.markdown("### 📁 Upload fichiers")
    st.caption("Optionnel — remplace les fichiers GitHub si uploadé")
    st.caption("Formats : `.xlsx` `.xls` `.csv`")
    file_ri = st.file_uploader("Randstad Intérim",
                                type=['xlsx','xls','csv'], key='ri',
                                help="Upload pour remplacer le fichier GitHub RI")
    file_exp = st.file_uploader("Expectra",
                                 type=['xlsx','xls','csv'], key='exp',
                                help="Upload pour remplacer le fichier GitHub EXP")

    # ── Vue fournisseur ───────────────────────────────────────────────
    has_ri  = file_ri  is not None or github_ri  is not None
    has_exp = file_exp is not None or github_exp is not None

    if has_ri or has_exp:
        st.markdown("---")
        st.markdown("### 🔎 Vue fournisseur")
        options_vue = []
        if has_ri and has_exp: options_vue.append('🔀 Consolidé RI+EXP')
        if has_ri:  options_vue.append('🏢 Randstad Intérim')
        if has_exp: options_vue.append('📊 Expectra')
        vue = st.radio("Vue fournisseur", options_vue, label_visibility='collapsed') if options_vue else None
    else:
        vue = None

    st.markdown("---")
    st.caption(f"📅 {date.today().strftime('%d/%m/%Y')}")

# ══════════════════════════════════════════════════════════════════════════
# CHARGEMENT DONNÉES
# ══════════════════════════════════════════════════════════════════════════
st.title("📊 Thales — Besoins & Candidatures")

aucune_source = not file_ri and not file_exp and not github_ri and not github_exp
if aucune_source:
    st.info("👈 **Choisissez une source de données** dans la barre latérale.")
    c1, c2 = st.columns(2)
    with c1:
        st.markdown("""
        #### ⬆️ Mode Upload
        Déposez vos fichiers directement dans l'app.
        - **Randstad Intérim** : pivot RI (20 colonnes)
        - **Expectra** : pivot EXP (21 colonnes)
        - **Formats** : `.xlsx` `.xls` `.csv`
        """)
    with c2:
        st.markdown("""
        #### 🔗 Mode GitHub
        Déposez vos fichiers dans le dossier `/data/`
        de votre repo GitHub — le dashboard se met
        à jour automatiquement sans action utilisateur.
        Configurez `GITHUB_RAW_URL` dans les secrets.
        """)
    st.stop()

@st.cache_data
def charger_ri(f):  return load_edb(f, idx_edb_hint=None)
@st.cache_data
def charger_exp(f): return load_edb(f, idx_edb_hint=11)

@st.cache_data(ttl=300)  # Cache 5 min pour GitHub
def charger_depuis_github(url, idx_edb_hint=None):
    """Charge un fichier depuis une URL GitHub raw via requests."""
    import requests, tempfile, os
    try:
        headers = {
            'User-Agent': 'Mozilla/5.0 (compatible; Streamlit)',
            'Accept': 'application/octet-stream, */*',
        }
        # Ajouter le token GitHub si disponible dans les secrets
        try:
            token = st.secrets.get("GITHUB_TOKEN", "")
            if token:
                headers["Authorization"] = f"token {token}"
        except Exception:
            pass

        r = requests.get(url, headers=headers, timeout=30)
        if r.status_code != 200:
            return None, f"HTTP Error {r.status_code}: {r.reason}"
        data = r.content
        ext = url.split('.')[-1].lower()
        with tempfile.NamedTemporaryFile(suffix=f'.{ext}', delete=False) as tmp:
            tmp.write(data)
            tmp_path = tmp.name
        class GithubFile:
            def __init__(self, path, name):
                self.name = name
                self._buf = io.BytesIO(open(path, 'rb').read())
            def read(self): return self._buf.read()
            def seek(self, n, w=0): return self._buf.seek(n, w)
            def tell(self): return self._buf.tell()
        gf = GithubFile(tmp_path, url.split('/')[-1])
        result = load_edb(gf, idx_edb_hint=idx_edb_hint)
        os.unlink(tmp_path)
        return result
    except Exception as e:
        return None, str(e)

df_ri = df_exp = df_conso = None

# ── Chargement RI ─────────────────────────────────────────────────────
if file_ri:
    try:
        df_ri = edb_to_df(charger_ri(file_ri), 'Randstad Intérim')
        st.sidebar.success(f"✅ RI (upload) — {len(df_ri)} expressions")
    except Exception as e:
        st.sidebar.error(f"❌ Erreur RI : {e}")
elif github_ri:
    with st.sidebar:
        with st.spinner("Chargement RI depuis GitHub..."):
            result = charger_depuis_github(github_ri, idx_edb_hint=None)
            if isinstance(result, tuple):
                st.error(f"❌ RI GitHub : {result[1]}")
            elif result:
                df_ri = edb_to_df(result, 'Randstad Intérim')
                st.success(f"✅ RI (GitHub) — {len(df_ri)} expressions")
            else:
                st.warning("⚠️ Fichier RI non trouvé sur GitHub")

# ── Chargement EXP ────────────────────────────────────────────────────
if file_exp:
    try:
        df_exp = edb_to_df(charger_exp(file_exp), 'Expectra')
        st.sidebar.success(f"✅ EXP (upload) — {len(df_exp)} expressions")
    except Exception as e:
        st.sidebar.error(f"❌ Erreur EXP : {e}")
elif github_exp:
    with st.sidebar:
        with st.spinner("Chargement EXP depuis GitHub..."):
            result = charger_depuis_github(github_exp, idx_edb_hint=11)
            if isinstance(result, tuple):
                st.error(f"❌ EXP GitHub : {result[1]}")
            elif result:
                df_exp = edb_to_df(result, 'Expectra')
                st.success(f"✅ EXP (GitHub) — {len(df_exp)} expressions")
            else:
                st.warning("⚠️ Fichier EXP non trouvé sur GitHub")

if df_ri is not None and df_exp is not None:
    df_conso = consolider(df_ri, df_exp)

# Vue active
if vue and 'Randstad' in vue and df_ri is not None:
    df_actif, nom_vue, couleur, has_statut = df_ri, "Randstad Intérim", "#1F6B75", False
elif vue and 'Expectra' in vue and df_exp is not None:
    df_actif, nom_vue, couleur, has_statut = df_exp, "Expectra", "#7030A0", True
elif df_conso is not None:
    df_actif, nom_vue, couleur, has_statut = df_conso, "Consolidé RI+EXP", "#1F3864", True
elif df_ri is not None:
    df_actif, nom_vue, couleur, has_statut = df_ri, "Randstad Intérim", "#1F6B75", False
elif df_exp is not None:
    df_actif, nom_vue, couleur, has_statut = df_exp, "Expectra", "#7030A0", True
else:
    df_actif = None
    nom_vue = ""; couleur = "#1F3864"; has_statut = False

if df_actif is None:
    st.warning("⚠️ Aucune donnée chargée. Uploadez un fichier ou vérifiez la connexion GitHub.")
    st.stop()

# ══════════════════════════════════════════════════════════════════════════
# NAVIGATION PAR ONGLETS
# ══════════════════════════════════════════════════════════════════════════
tab_db, tab_rech, tab_stats, tab_crit = st.tabs([
    "📊 Tableau de Bord",
    "🔍 Recherche",
    "📈 Analyses & Statistiques",
    "🚨 Actions Requises",
])

# ══════════════════════════════════════════════════════════════════════════
# ONGLET 1 : TABLEAU DE BORD
# ══════════════════════════════════════════════════════════════════════════
with tab_db:
    # Filtres rapides
    with st.expander("🔎 Filtres", expanded=False):
        fc1, fc2, fc3, fc4 = st.columns(4)
        with fc1:
            sites_all = sorted(df_actif['site'].dropna().unique())
            site_sel = st.multiselect("Site Thales", sites_all)
        with fc2:
            sems_all = sorted(df_actif['sem'].dropna().unique())
            sem_sel = st.multiselect("Semaine diffusion", sems_all)
        with fc3:
            if has_statut:
                statuts_all = sorted(df_actif['statut_edb'].dropna().unique())
                statut_sel = st.multiselect("Statut EdB", statuts_all)
            else:
                statut_sel = []
        with fc4:
            divs_all = sorted(df_actif['division'].dropna().unique())
            div_sel = st.multiselect("Division Thales", divs_all)

    df_f = df_actif.copy()
    if site_sel:   df_f = df_f[df_f['site'].isin(site_sel)]
    if sem_sel:    df_f = df_f[df_f['sem'].isin(sem_sel)]
    if statut_sel: df_f = df_f[df_f['statut_edb'].isin(statut_sel)]
    if div_sel:    df_f = df_f[df_f['division'].isin(div_sel)]

    st.markdown(f"**Vue : {nom_vue}** — `{len(df_f)}` expressions"
                + (f" *(filtres actifs)*" if len(df_f) < len(df_actif) else ""))
    st.markdown("<br>", unsafe_allow_html=True)

    kpi_row(df_f)
    st.markdown("<br>", unsafe_allow_html=True)

    c1, c2 = st.columns([3, 2])
    with c1:
        st.plotly_chart(fig_barres_semaine(df_f, f"Expressions par Semaine — {nom_vue}"),
                        use_container_width=True)
    with c2:
        st.plotly_chart(fig_camembert(df_f, f"Répartition — {nom_vue}"),
                        use_container_width=True)

    st.plotly_chart(fig_sites(df_f, f"Top 15 Sites en tension — {nom_vue}"),
                    use_container_width=True)

    # Comparaison RI vs EXP si consolidé
    if nom_vue == "Consolidé RI+EXP" and df_ri is not None and df_exp is not None:
        st.markdown("---")
        st.markdown('<p class="section-title">🔄 Comparaison Randstad Intérim vs Expectra</p>', unsafe_allow_html=True)
        c1, c2 = st.columns(2)
        with c1: st.plotly_chart(fig_camembert(df_ri, "Randstad Intérim"), use_container_width=True)
        with c2: st.plotly_chart(fig_camembert(df_exp, "Expectra"), use_container_width=True)

    st.markdown("---")
    export_excel(df_f, f"Thales_{nom_vue.replace(' ','_')}")

# ══════════════════════════════════════════════════════════════════════════
# ONGLET 2 : RECHERCHE
# ══════════════════════════════════════════════════════════════════════════
with tab_rech:
    st.markdown('<p class="section-title">🔍 Recherche multicritère</p>', unsafe_allow_html=True)

    with st.container():
        r1c1, r1c2, r1c3 = st.columns(3)
        with r1c1:
            q_texte = st.text_input("🔤 Recherche libre",
                placeholder="N° expression, qualification, site...",
                help="Recherche dans tous les champs texte")
        with r1c2:
            q_sem_range = st.text_input("📅 Semaine(s)",
                placeholder="ex: 06 2026 ou 06 2026 - 10 2026",
                help="Une semaine ou une plage")
        with r1c3:
            pass  # espaceur

        # Ligne SIRET : liste déroulante + saisie libre côte à côte
        st.markdown("**🏢 Recherche par SIRET**")
        s1, s2, s3 = st.columns([2, 2, 2])
        with s1:
            # Liste déroulante des SIRET connus
            sirets_connus = sorted(
                [s for s in df_actif['siret'].dropna().unique() if s.strip() and s != 'None'],
                key=lambda x: x
            )
            # Enrichir avec le nom du site pour faciliter la lecture
            siret_to_site = df_actif.drop_duplicates('siret').set_index('siret')['site'].to_dict()
            siret_labels = ['Tous'] + [
                f"{s}  —  {siret_to_site.get(s, '')[:30]}" for s in sirets_connus
            ]
            siret_vals = [''] + sirets_connus
            siret_idx = st.selectbox(
                "Sélectionner un SIRET",
                range(len(siret_labels)),
                format_func=lambda i: siret_labels[i],
                key='siret_select'
            )
            q_siret_select = siret_vals[siret_idx]
        with s2:
            q_siret_libre = st.text_input(
                "Ou saisir librement",
                placeholder="ex: 41472510100086",
                help="Saisie partielle acceptée",
                key='siret_libre'
            )
        with s3:
            if q_siret_select:
                st.info(f"📌 Site : **{siret_to_site.get(q_siret_select, '—')}**")
            elif q_siret_libre.strip():
                # Trouver le site correspondant
                match = df_actif[df_actif['siret'].astype(str).str.contains(q_siret_libre.strip(), na=False)]
                if not match.empty:
                    sites_trouves = match['site'].dropna().unique()
                    st.info(f"📌 Site(s) : **{', '.join(sites_trouves[:3])}**")

        # SIRET actif = liste OU saisie libre (priorité à la liste si les deux sont remplis)
        q_siret_actif = q_siret_select if q_siret_select else q_siret_libre.strip()

        r2c1, r2c2, r2c3, r2c4 = st.columns(4)
        with r2c1:
            sites_r = ['Tous'] + sorted(df_actif['site'].dropna().unique())
            q_site = st.selectbox("📍 Site Thales", sites_r)
        with r2c2:
            divs_r = ['Toutes'] + sorted(df_actif['division'].dropna().unique())
            q_div = st.selectbox("🏭 Division Thales", divs_r)
        with r2c3:
            quals_r = ['Toutes'] + sorted(df_actif['qual'].dropna().unique())
            q_qual = st.selectbox("🎓 Qualification", quals_r)
        with r2c4:
            cats_r = ['Toutes'] + [CAT_LBL[c] for c in CAT_ORDER]
            q_cat = st.selectbox("📋 Situation", cats_r)

    # Appliquer la recherche
    df_r = df_actif.copy()

    if q_texte.strip():
        mask = False
        for col in ['num', 'site', 'qual', 'agences', 'siret']:
            if col in df_r.columns:
                mask = mask | df_r[col].astype(str).str.contains(q_texte.strip(), case=False, na=False)
        df_r = df_r[mask]

    if q_siret_actif:
        if q_siret_select:
            # Correspondance exacte si sélection depuis la liste
            df_r = df_r[df_r['siret'].astype(str) == q_siret_actif]
        else:
            # Correspondance partielle si saisie libre
            df_r = df_r[df_r['siret'].astype(str).str.contains(q_siret_actif, na=False)]

    if q_sem_range.strip():
        parts = [p.strip() for p in q_sem_range.split('-') if p.strip()]
        if len(parts) == 2:
            sem_min, sem_max = parts[0], parts[1]
            df_r = df_r[(df_r['sem'] >= sem_min) & (df_r['sem'] <= sem_max)]
        elif len(parts) == 1:
            df_r = df_r[df_r['sem'] == parts[0]]

    if q_site != 'Tous':     df_r = df_r[df_r['site'] == q_site]
    if q_div  != 'Toutes':   df_r = df_r[df_r['division'] == q_div]
    if q_qual != 'Toutes':   df_r = df_r[df_r['qual'] == q_qual]
    if q_cat  != 'Toutes':
        cat_key = {v:k for k,v in CAT_LBL.items()}.get(q_cat)
        if cat_key: df_r = df_r[df_r['cat'] == cat_key]

    # Résultats
    st.markdown(f"**{len(df_r)} résultat(s)** trouvé(s)")

    if not df_r.empty:
        # KPIs résultats
        kpi_row(df_r)
        st.markdown("<br>", unsafe_allow_html=True)

        # Tableau résultats
        cols_show = ['num','site','division','qual','sem','statut_edb','cat_label','nb_cand','siret']
        cols_show = [c for c in cols_show if c in df_r.columns]
        rename_r = {'num':'N° Expression','site':'Site','division':'Division',
                    'qual':'Qualification','sem':'Semaine','statut_edb':'Statut EdB',
                    'cat_label':'Situation','nb_cand':'Nb Cand.','siret':'SIRET'}

        def color_row(row):
            s = row.get('Situation', '')
            if 'Sans candidature' in str(s): return ['background-color:#FCE4D6']*len(row)
            if 'refusées' in str(s).lower():  return ['background-color:#FBE4D5']*len(row)
            if 'Acceptée' in str(s):           return ['background-color:#E2EFDA']*len(row)
            if 'sélectionner' in str(s) or 'étudier' in str(s): return ['background-color:#D6E4F0']*len(row)
            return ['']*len(row)

        df_show = df_r[cols_show].rename(columns=rename_r)
        st.dataframe(df_show.style.apply(color_row, axis=1),
                     use_container_width=True, hide_index=True,
                     height=min(500, 45 + len(df_show)*36))

        # Mini graphiques résultats
        if len(df_r) > 1:
            rc1, rc2 = st.columns(2)
            with rc1:
                st.plotly_chart(fig_camembert(df_r, "Répartition résultats"),
                                use_container_width=True)
            with rc2:
                if len(df_r['site'].unique()) > 1:
                    st.plotly_chart(fig_sites(df_r, "Sites (résultats)", top_n=10),
                                    use_container_width=True)

        st.markdown("---")
        export_excel(df_r, "Thales_Recherche")
    else:
        st.warning("Aucun résultat pour cette recherche.")

# ══════════════════════════════════════════════════════════════════════════
# ONGLET 3 : ANALYSES & STATISTIQUES
# ══════════════════════════════════════════════════════════════════════════
with tab_stats:
    st.markdown('<p class="section-title">📈 Analyses & Statistiques — ' + nom_vue + '</p>',
                unsafe_allow_html=True)

    # Filtre rapide sur statut pour les analyses
    if has_statut:
        col_f1, col_f2 = st.columns([2,4])
        with col_f1:
            statuts_s = ['Tous'] + sorted(df_actif['statut_edb'].dropna().unique())
            st_sel = st.selectbox("Statut EdB", statuts_s, key='stat_statut')
        df_s = df_actif[df_actif['statut_edb'] == st_sel] if st_sel != 'Tous' else df_actif.copy()
    else:
        df_s = df_actif.copy()

    # ── 1. Taux de couverture ─────────────────────────────────────────────
    st.markdown('<p class="section-title">1. Taux de couverture par site</p>', unsafe_allow_html=True)
    c1, c2 = st.columns(2)
    with c1:
        st.plotly_chart(fig_taux_couverture(df_s, 'site', "Taux de couverture par Site Thales"),
                        use_container_width=True)
    with c2:
        st.plotly_chart(fig_taux_couverture(df_s, 'division', "Taux de couverture par Division Thales", top_n=10),
                        use_container_width=True)

    # ── 2. Évolution temporelle ───────────────────────────────────────────
    st.markdown('<p class="section-title">2. Évolution temporelle</p>', unsafe_allow_html=True)
    st.plotly_chart(fig_tendance(df_s, f"Évolution semaine par semaine — {nom_vue}"),
                    use_container_width=True)

    # ── 3. Qualifications en tension ─────────────────────────────────────
    st.markdown('<p class="section-title">3. Qualifications en tension</p>', unsafe_allow_html=True)
    top_n_q = st.slider("Nombre de qualifications à afficher", 10, 40, 20, key='slider_qual')
    st.plotly_chart(fig_qualifications(df_s, top_n=top_n_q), use_container_width=True)

    # ── 4. Délais moyens ─────────────────────────────────────────────────
    st.markdown('<p class="section-title">4. Délais Diffusion → Début mission</p>', unsafe_allow_html=True)
    df_delai = df_s[df_s['delai'].notna()]
    if not df_delai.empty:
        # Stats globales délais
        d1, d2, d3, d4 = st.columns(4)
        with d1:
            st.markdown(f"""<div class="metric-card blue">
                <p>Délai moyen</p><h2>{df_delai['delai'].mean():.1f}j</h2></div>""",
                unsafe_allow_html=True)
        with d2:
            st.markdown(f"""<div class="metric-card">
                <p>Médiane</p><h2>{df_delai['delai'].median():.1f}j</h2></div>""",
                unsafe_allow_html=True)
        with d3:
            pos = df_delai[df_delai['delai'] > 0]
            st.markdown(f"""<div class="metric-card green">
                <p>Anticipées (>0j)</p><h2>{len(pos)}</h2>
                <p class="pct">{len(pos)/len(df_delai)*100:.0f}%</p></div>""",
                unsafe_allow_html=True)
        with d4:
            retro = df_delai[df_delai['delai'] <= 0]
            st.markdown(f"""<div class="metric-card orange">
                <p>Rétroactives (≤0j)</p><h2>{len(retro)}</h2>
                <p class="pct">{len(retro)/len(df_delai)*100:.0f}%</p></div>""",
                unsafe_allow_html=True)
        st.markdown("<br>", unsafe_allow_html=True)
        fig_d = fig_delais(df_s, 'site', "Délais moyens par Site Thales")
        if fig_d: st.plotly_chart(fig_d, use_container_width=True)
    else:
        st.info("Aucune donnée de délai disponible dans ce fichier.")

    # ── 5. Comparaison RI vs EXP par site ────────────────────────────────
    if nom_vue == "Consolidé RI+EXP" and df_ri is not None and df_exp is not None:
        st.markdown('<p class="section-title">5. Performance RI vs EXP par site</p>', unsafe_allow_html=True)
        fig_perf = fig_ri_vs_exp(df_ri, df_exp)
        if fig_perf:
            st.plotly_chart(fig_perf, use_container_width=True)
            st.caption("Chaque bulle = un site commun RI+EXP. Taille proportionnelle au nombre d'expressions.")

    # ── 6. Tableau statistiques récapitulatif ────────────────────────────
    st.markdown('<p class="section-title">6. Tableau statistiques par site</p>', unsafe_allow_html=True)
    grp_stat = df_s.groupby('site').apply(lambda x: pd.Series({
        'Total': len(x),
        '✅ Acceptées': (x['cat']=='acceptee').sum(),
        '🔵 En cours': ((x['cat']=='a_selectionner')|(x['cat']=='a_etudier')).sum(),
        '🟠 Refusées': (x['cat']=='toutes_refusees').sum(),
        '🔴 Sans cand.': (x['cat']=='sans_cand').sum(),
        'Taux couverture': f"{(x['cat']!='sans_cand').sum()/len(x)*100:.0f}%",
        'Délai moyen (j)': f"{x['delai'].mean():.1f}" if x['delai'].notna().any() else '—',
    })).reset_index().sort_values('🔴 Sans cand.', ascending=False)

    st.dataframe(grp_stat, use_container_width=True, hide_index=True,
                 height=min(500, 45 + len(grp_stat)*36))
    st.markdown("---")
    export_excel(df_s, f"Thales_Stats_{nom_vue.replace(' ','_')}")

# ══════════════════════════════════════════════════════════════════════════
# ONGLET 4 : ACTIONS REQUISES
# ══════════════════════════════════════════════════════════════════════════
with tab_crit:
    st.markdown('<p class="section-title">🚨 Expressions nécessitant une action immédiate</p>',
                unsafe_allow_html=True)

    df_diff = df_actif[df_actif['statut_edb'] == 'Diffusée'] if has_statut else df_actif

    tab_s, tab_r = st.tabs([
        f"🔴 Sans candidature ({(df_diff['cat']=='sans_cand').sum()})",
        f"🟠 Toutes refusées ({(df_diff['cat']=='toutes_refusees').sum()})",
    ])

    cols_crit = ['num','site','division','qual','sem','cat_label','nb_cand','siret','agences']
    cols_crit = [c for c in cols_crit if c in df_diff.columns]
    rename_crit = {'num':'N° Expression','site':'Site','division':'Division',
                   'qual':'Qualification','sem':'Semaine','cat_label':'Situation',
                   'nb_cand':'Nb Cand.','siret':'SIRET','agences':'Agences'}

    for tab_obj, cat_id, bg_color in [
        (tab_s, 'sans_cand',      '#FCE4D6'),
        (tab_r, 'toutes_refusees','#FBE4D5'),
    ]:
        with tab_obj:
            df_c = df_diff[df_diff['cat'] == cat_id].sort_values('site')
            if df_c.empty:
                st.success("✅ Aucune expression dans cette catégorie.")
            else:
                # Filtre rapide dans l'onglet
                fc1, fc2 = st.columns(2)
                with fc1:
                    sites_c = ['Tous'] + sorted(df_c['site'].dropna().unique())
                    site_c = st.selectbox("Filtrer par site", sites_c, key=f'site_{cat_id}')
                with fc2:
                    divs_c = ['Toutes'] + sorted(df_c['division'].dropna().unique())
                    div_c = st.selectbox("Filtrer par division", divs_c, key=f'div_{cat_id}')

                if site_c != 'Tous':   df_c = df_c[df_c['site'] == site_c]
                if div_c  != 'Toutes': df_c = df_c[df_c['division'] == div_c]

                st.markdown(f"**{len(df_c)} expression(s)**")
                df_show_c = df_c[cols_crit].rename(columns=rename_crit)
                st.dataframe(
                    df_show_c.style.apply(lambda _: [f'background-color:{bg_color}']*len(_), axis=1),
                    use_container_width=True, hide_index=True,
                    height=min(500, 45 + len(df_show_c)*36)
                )
                st.markdown("---")
                export_excel(df_c, f"Thales_Critiques_{cat_id}")
